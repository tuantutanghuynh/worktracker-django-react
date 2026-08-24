from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.core.cache import cache

DASHBOARD_CACHE_KEY = 'admin:dashboard'
DASHBOARD_CACHE_TTL = 30  # seconds — dashboard data cũ tối đa 30 giây

from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.views import APIView
from rest_framework.response import Response

from accounts.models import CustomUser, Department
from accounts.permissions import HasPermission
from system.security.permissions_manager import IsAdminRole
from system.pagination import AdminPageNumberPagination
from projects.models import Client, Job
from ..models import AuditLog
from .serializers import AuditLogSerializer
from django.http import HttpResponse
from ..utils import log_audit_event
import openpyxl

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = AuditLogSerializer
    pagination_class = AdminPageNumberPagination
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['created_at', 'user__email', 'action', 'table_name', 'record_id', 'severity']

    def get_permissions(self):
        return [IsAdminRole(), HasPermission('audit:view')]

    def get_queryset(self):
        queryset = AuditLog.objects.all().order_by('-created_at')

        actor = self.request.query_params.get('actor')
        if actor:
            queryset = queryset.filter(user_id=actor)

        if actor_role := self.request.query_params.get('actor_role'):
            queryset = queryset.filter(user__role__code=actor_role)

        action = self.request.query_params.get('action')
        if action:
            queryset = queryset.filter(action=action)

        table_name = self.request.query_params.get('table_name')
        if table_name:
            queryset = queryset.filter(table_name=table_name)

        date_from = self.request.query_params.get('date_from')
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)

        date_to = self.request.query_params.get('date_to')
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        if record_id := self.request.query_params.get('record_id'):
            queryset = queryset.filter(record_id=record_id)

        if severity := self.request.query_params.get('severity'):
            queryset = queryset.filter(severity=severity)

        if keyword := self.request.query_params.get('keyword'):
            # 'create' phải khớp cả dòng action=CREATE, không chỉ tìm chữ
            # "create" nằm lẫn đâu đó trong nội dung old/new_values.
            queryset = queryset.filter(
                Q(old_values__icontains=keyword)
                | Q(new_values__icontains=keyword)
                | Q(action__icontains=keyword)
                | Q(table_name__icontains=keyword)
            )

        return queryset

    # Powers the Action/Table filter dropdowns on the frontend with the
    # values actually present in the table, instead of a hardcoded list
    # that would drift as new action types get added across the app.
    @action(detail=False, methods=["get"], url_path="filters")
    def filter_options(self, request):
        return Response({
            'actions': list(
                AuditLog.objects.order_by('action').values_list('action', flat=True).distinct()
            ),
            'tables': list(
                AuditLog.objects.order_by('table_name').values_list('table_name', flat=True).distinct()
            ),
        })

    # Powers the 5 KPI cards at the top of the Audit Logs page. Scoped to
    # today + whichever role tab is active (?actor_role=), same scoping as
    # the table itself, so the cards always describe what's on screen.
    @action(detail=False, methods=["get"], url_path="summary")
    def summary_stats(self, request):
        today = timezone.now().date()
        qs = AuditLog.objects.filter(created_at__date=today)

        if actor_role := request.query_params.get('actor_role'):
            qs = qs.filter(user__role__code=actor_role)

        return Response({
            'total_logs_today': qs.count(),
            'sensitive_actions': qs.filter(
                severity__in=[AuditLog.Severity.CRITICAL, AuditLog.Severity.WARNING]
            ).count(),
            'account_changes': qs.filter(table_name='users').count(),
            'timesheet_locks': qs.filter(action__in=['LOCK_TIMESHEET', 'UNLOCK_TIMESHEET']).count(),
            'data_changes': qs.filter(
                table_name__in=['jobs', 'clients', 'tasks'], action__in=['CREATE', 'UPDATE']
            ).count(),
        })


class DashboardView(APIView):

    def get_permissions(self):
        return [IsAdminRole(), HasPermission('audit:view')]

    def get(self, request):
        cached = cache.get(DASHBOARD_CACHE_KEY)
        if cached:
            return Response(cached)

        today = timezone.now().date()

        active_clients = Client.objects.filter(is_active=True).count()
        total_users    = CustomUser.objects.count()
        active_accounts = CustomUser.objects.filter(is_active=True).count()
        locked_accounts = CustomUser.objects.filter(is_active=False).count()
        departments_without_manager = Department.objects.filter(manager__isnull=True).count()
        overdue_jobs   = Job.objects.filter(
            deadline__lt=today,
            status__in=['PLANNING', 'ACTIVE', 'ON_HOLD'],
        ).count()

        jobs_by_status = {
            s: Job.objects.filter(status=s).count()
            for s in ['PLANNING', 'ACTIVE', 'ON_HOLD', 'COMPLETED', 'CANCELLED']
        }
        jobs_by_status['OVERDUE'] = overdue_jobs

        clients_overview = {
            'active':   active_clients,
            'inactive': Client.objects.filter(is_active=False).count(),
            'total':    Client.objects.count(),
        }

        # IAM/security activity — table_name kept explicit alongside action
        # so this stays correct even if the same action name is ever reused
        # against a different table_name elsewhere.
        audit_today = AuditLog.objects.filter(created_at__date=today)
        audit_summary_today = {
            'account_created':  audit_today.filter(action='CREATE', table_name='users').count(),
            'account_locked':   audit_today.filter(action='LOCK_ACCOUNT', table_name='users').count(),
            'role_changed':     audit_today.filter(action='ROLE_CHANGED', table_name='users').count(),
            'password_reset':   audit_today.filter(action='RESET_PASSWORD', table_name='users').count(),
        }

        # Quick-glance feed of the most recent sensitive IAM actions — mirrors
        # the severity now set on ROLE_CHANGED/LOCK_ACCOUNT/UNLOCK_ACCOUNT/
        # RESET_PASSWORD in accounts/admin/views.py.
        recent_security_events = AuditLog.objects.filter(
            severity__in=[AuditLog.Severity.CRITICAL, AuditLog.Severity.WARNING]
        ).order_by('-created_at')[:5]

        data = {
            'active_clients':               active_clients,
            'total_users':                  total_users,
            'active_accounts':              active_accounts,
            'locked_accounts':              locked_accounts,
            'departments_without_manager':  departments_without_manager,
            'jobs_by_status':               jobs_by_status,
            'clients_overview':             clients_overview,
            'audit_summary_today':          audit_summary_today,
            'recent_security_events':       AuditLogSerializer(recent_security_events, many=True).data,
        }
        cache.set(DASHBOARD_CACHE_KEY, data, timeout=DASHBOARD_CACHE_TTL)
        return Response(data)

class DataQualityAlertsView(APIView):
    """
    Synthetic, not-persisted alerts computed live from current DB state —
    unlike AuditLog-triggered notifications, these have no "read/unread"
    concept: they simply stop appearing once the underlying field is fixed,
    with no stale row to clean up afterwards.
    """

    def get_permissions(self):
        return [IsAdminRole(), HasPermission('audit:view')]

    def get(self, request):
        alerts = []

        for dept in Department.objects.filter(manager__isnull=True):
            alerts.append({
                'id': f'dept-no-manager-{dept.id}',
                'title': 'Department has no manager',
                'content': f'"{dept.name}" has no manager assigned.',
                'related_url': f'/admin/departments?edit={dept.id}',
            })

        employees_without_department = CustomUser.objects.filter(
            role__code='EMPLOYEE', is_active=True
        ).filter(Q(profile__isnull=True) | Q(profile__department__isnull=True))
        for user in employees_without_department:
            alerts.append({
                'id': f'user-no-department-{user.id}',
                'title': 'Employee not assigned to a department',
                'content': f'{user.email} has no department.',
                'related_url': f'/admin/users/search?edit={user.id}',
            })

        clients_missing_contact = Client.objects.filter(is_active=True).filter(
            Q(contact_email__isnull=True) | Q(contact_email='')
            | Q(contact_phone__isnull=True) | Q(contact_phone='')
        )
        for c in clients_missing_contact:
            alerts.append({
                'id': f'client-missing-contact-{c.id}',
                'title': 'Client missing contact info',
                'content': f'"{c.client_name}" has no contact email or phone.',
                'related_url': f'/admin/clients?edit={c.id}',
            })

        return Response(alerts)


class AdminReportView(APIView):
    def get_permissions(self):
        return [IsAdminRole(), HasPermission('report:export')]

    @transaction.atomic
    def get(self, request):
        # Filter theo khoảng thời gian tạo job (optional)
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clients'
        ws.append(['ID', 'Name', 'Tax Code', 'Contact Email', 'Is Active'])
        for c in Client.objects.all():
            ws.append([c.id, c.client_name, c.tax_code, c.contact_email, c.is_active])

        ws2 = wb.create_sheet('Jobs')
        ws2.append(['ID', 'Name', 'Client', 'Status', 'Priority', 'Start Date', 'Deadline'])
        job_qs = Job.objects.select_related('client').all()
        if date_from:
            job_qs = job_qs.filter(created_at__date__gte=date_from)
        if date_to:
            job_qs = job_qs.filter(created_at__date__lte=date_to)
        for j in job_qs:
            ws2.append([j.id, j.job_name, j.client.client_name, j.status, j.priority,
                        str(j.start_date), str(j.deadline)])
        #tạo sheet user
        ws3 = wb.create_sheet('Users')
        ws3.append(['ID', 'Email', 'Role', 'Is Active'])
        for u in CustomUser.objects.select_related('role').all():
            ws3.append([u.id, u.email, u.role.code if u.role else '', u.is_active])            
        log_audit_event(
            actor=request.user,
            action='EXPORT',
            table_name='reports',
            record_id=0,
            request=request,
        )
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="worktracker_report.xlsx"'
        wb.save(response)
        return response
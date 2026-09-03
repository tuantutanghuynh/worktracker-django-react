"""
Module: system.services.admin_report_export_service
Description: Excel spreadsheet generation and row formatting utilities for system administration reports.
"""

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY_HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
REGULAR_FONT = Font(name="Calibri", size=10, color="1E293B")

THIN_BORDER_SIDE = Side(border_style="thin", color="CBD5E1")
THIN_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def write_sheet(sheet, headers, rows):
    """Write styled headers and zebra-striped rows into openpyxl worksheet."""
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = NAVY_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for row in rows:
        sheet.append(row)

    for r in range(2, sheet.max_row + 1):
        row_fill = ZEBRA_FILL if r % 2 == 0 else WHITE_FILL
        for c in range(1, len(headers) + 1):
            cell = sheet.cell(row=r, column=c)
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            cell.fill = row_fill

    sheet.freeze_panes = "A2"

    for col in sheet.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = max((len(str(cell.value or "")) for cell in col), default=0)
        sheet.column_dimensions[col_letter].width = min(max(max_length + 3, 11), 45)


def build_xlsx_response(*, sheet_title, headers, rows, filename):
    """Generate and return an HTTP response streaming an Excel workbook attachment."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
    write_sheet(sheet, headers, rows)

    response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


CLIENT_HEADERS = ["ID", "Client Name", "Tax Code", "Contact Person", "Contact Email", "Contact Phone", "Status", "Created At"]


def client_rows(queryset):
    """Map client queryset into Excel row values list."""
    return [
        [
            c.id, c.client_name, c.tax_code, c.contact_person or "", c.contact_email or "",
            c.contact_phone or "", "Active" if c.is_active else "Inactive",
            c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
        ]
        for c in queryset
    ]


JOB_HEADERS = ["ID", "Job Code", "Job Name", "Client", "Manager", "Priority", "Status", "Start Date", "Deadline"]


def job_rows(queryset):
    """Map job queryset into Excel row values list."""
    return [
        [
            j.id, j.job_code or "", j.job_name, j.client.client_name, j.manager.email,
            j.priority, j.status, str(j.start_date), str(j.deadline),
        ]
        for j in queryset.select_related("client", "manager")
    ]


USER_HEADERS = ["ID", "Email", "Full Name", "Role", "Department", "Status"]


def user_rows(queryset):
    """Map user account queryset into Excel row values list."""
    return [
        [
            u.id, u.email,
            getattr(getattr(u, "profile", None), "full_name", "") or "",
            u.role.name if u.role else "",
            getattr(getattr(getattr(u, "profile", None), "department", None), "name", "") or "",
            "Active" if u.is_active else "Locked",
        ]
        for u in queryset.select_related("role", "profile", "profile__department")
    ]


DEPARTMENT_HEADERS = ["ID", "Name", "Description", "Manager", "Created At"]


def department_rows(queryset):
    """Map department queryset into Excel row values list."""
    return [
        [
            d.id, d.name, d.description or "",
            d.manager.email if d.manager else "",
            d.created_at.strftime("%Y-%m-%d") if d.created_at else "",
        ]
        for d in queryset.select_related("manager")
    ]


AUDIT_LOG_HEADERS = ["ID", "Time", "Actor", "Action", "Module", "Record ID", "Severity", "Summary", "IP Address"]


def audit_log_rows(queryset):
    """Map audit log queryset into Excel row values list."""
    return [
        [
            a.id,
            a.created_at.strftime("%Y-%m-%d %H:%M:%S") if a.created_at else "",
            a.user.email if a.user else "",
            a.action, a.table_name, a.record_id, a.severity,
            a.summary or "", a.ip_address or "",
        ]
        for a in queryset.select_related("user")
    ]


TIMESHEET_HEADERS = ["Employee", "Email", "Department", "Month Hours", "Target Hours", "Avg/Day", "Violations", "Missing Days", "Status", "Last Entry"]


def timesheet_rows(rows):
    """Map pre-aggregated timesheet summary records into Excel row values list."""
    return [
        [
            row["full_name"], row["email"], row["department_name"] or "",
            row["month_hours"], row["target_hours"], row["avg_per_day"],
            row["violations"], row["missing_days"], row["status"],
            str(row["last_entry"] or ""),
        ]
        for row in rows
    ]

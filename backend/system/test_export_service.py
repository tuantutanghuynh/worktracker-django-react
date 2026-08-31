"""
Test xuất báo cáo Excel của Admin.

Hai nhóm rủi ro khác nhau, test tách riêng:

  1. build_xlsx_response  — file tạo ra có mở được không, có đúng số dòng
     không. Lỗi ở đây làm hỏng file, Excel báo "file bị hư" khi mở.

  2. Endpoint export      — có tôn trọng bộ lọc đang xem không. Đây mới là
     lỗi nguy hiểm: file vẫn mở được bình thường nhưng chứa dữ liệu sai —
     xuất ra toàn bộ bảng trong khi người dùng đang lọc, hoặc ngược lại.
     Không ai phát hiện ra cho tới khi đã gửi báo cáo đi.
"""
from io import BytesIO

import pytest
from model_bakery import baker
from openpyxl import load_workbook
from rest_framework.test import APIClient

from accounts.models import Permission, RolePermission
from projects.models import Client as ClientModel
from system.services import admin_report_export_service as svc


XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def doc_workbook(response):
    """Mở lại file vừa xuất — cách duy nhất chứng minh nó thật sự hợp lệ."""
    return load_workbook(BytesIO(response.content))


class TestBuildXlsxResponse:

    def test_tra_ve_dung_content_type_va_ten_file(self):
        response = svc.build_xlsx_response(
            sheet_title="Thu", headers=["A", "B"], rows=[[1, 2]], filename="thu.xlsx"
        )
        assert response["Content-Type"] == XLSX_CONTENT_TYPE
        assert 'attachment; filename="thu.xlsx"' in response["Content-Disposition"]

    def test_file_mo_duoc_va_du_dong(self):
        rows = [[i, f"ten-{i}"] for i in range(1, 6)]
        response = svc.build_xlsx_response(
            sheet_title="Danh sach", headers=["ID", "Ten"], rows=rows, filename="x.xlsx"
        )
        sheet = doc_workbook(response).active

        assert sheet.title == "Danh sach"
        # 1 dòng tiêu đề + 5 dòng dữ liệu
        assert sheet.max_row == 6
        assert [c.value for c in sheet[1]] == ["ID", "Ten"]
        assert sheet.cell(row=2, column=2).value == "ten-1"

    def test_khong_co_dong_nao_van_ra_file_hop_le(self):
        """
        Lọc ra 0 kết quả rồi bấm Export là chuyện thường. Không được ném lỗi
        hay trả file hỏng — phải là file chỉ có dòng tiêu đề.
        """
        response = svc.build_xlsx_response(
            sheet_title="Rong", headers=["ID", "Ten"], rows=[], filename="rong.xlsx"
        )
        sheet = doc_workbook(response).active
        assert sheet.max_row == 1
        assert [c.value for c in sheet[1]] == ["ID", "Ten"]


@pytest.mark.django_db
class TestRowMappers:

    def test_client_rows_khop_so_cot_voi_header(self):
        baker.make("projects.Client", client_name="Cong ty A", is_active=True)
        rows = svc.client_rows(ClientModel.objects.all())

        assert len(rows) == 1
        assert len(rows[0]) == len(svc.CLIENT_HEADERS), (
            "So cot du lieu phai khop header, lech la file xuat ra lech cot"
        )

    def test_client_rows_doi_is_active_thanh_chu(self):
        baker.make("projects.Client", client_name="B", is_active=False)
        rows = svc.client_rows(ClientModel.objects.all())
        assert rows[0][-2] == "Inactive"

    def test_truong_rong_thanh_chuoi_rong_khong_phai_None(self):
        """
        openpyxl ghi None thành ô trống, nhưng chuỗi "None" thì hiện ra chữ
        "None" trong file — xấu và gây hiểu nhầm. Mapper phải đổi sẵn.
        """
        baker.make(
            "projects.Client", client_name="C", contact_person=None,
            contact_email=None, contact_phone=None,
        )
        rows = svc.client_rows(ClientModel.objects.all())
        assert rows[0][3] == ""
        assert rows[0][4] == ""


@pytest.mark.django_db
class TestExportEndpointTonTrongBoLoc:
    """
    Luật quan trọng nhất của tính năng Export: file xuất ra phải chứa ĐÚNG
    những dòng người dùng đang nhìn thấy, không nhiều hơn.

    Endpoint dùng lại filter_queryset(get_queryset()) nên về nguyên tắc là
    tự động đúng — nhưng đó chính là loại giả định cần một test canh giữ,
    vì chỉ cần ai đó viết lại get_queryset là hỏng mà không ai biết.
    """

    @pytest.fixture
    def admin_client(self, db):
        role = baker.make("accounts.Role", code="ADMIN")
        for code in ("client:view", "client:export"):
            perm, _ = Permission.objects.get_or_create(code=code, defaults={"name": code})
            RolePermission.objects.get_or_create(role=role, permission=perm)
        admin = baker.make(
            "accounts.CustomUser", role=role, is_active=True, must_change_password=False
        )
        c = APIClient()
        c.force_authenticate(user=admin)
        return c

    @pytest.fixture
    def du_lieu(self, db):
        baker.make("projects.Client", client_name="Alpha Corp", is_active=True)
        baker.make("projects.Client", client_name="Beta Ltd", is_active=True)
        baker.make("projects.Client", client_name="Gamma Inc", is_active=False)

    def test_khong_loc_thi_xuat_het(self, admin_client, du_lieu):
        response = admin_client.get("/api/admin/clients/export/")
        assert response.status_code == 200
        sheet = doc_workbook(response).active
        assert sheet.max_row == 4  # 1 tieu de + 3 client

    def test_loc_theo_trang_thai_thi_chi_xuat_dong_khop(self, admin_client, du_lieu):
        response = admin_client.get("/api/admin/clients/export/?is_active=false")
        sheet = doc_workbook(response).active

        assert sheet.max_row == 2  # 1 tieu de + 1 client
        assert sheet.cell(row=2, column=2).value == "Gamma Inc"

    def test_loc_theo_ten_thi_chi_xuat_dong_khop(self, admin_client, du_lieu):
        response = admin_client.get("/api/admin/clients/export/?search=Alpha")
        sheet = doc_workbook(response).active

        assert sheet.max_row == 2
        assert sheet.cell(row=2, column=2).value == "Alpha Corp"

    def test_loc_khong_ra_gi_thi_file_chi_co_tieu_de(self, admin_client, du_lieu):
        response = admin_client.get("/api/admin/clients/export/?search=KhongTonTai")
        sheet = doc_workbook(response).active
        assert sheet.max_row == 1

    def test_export_khong_bi_gioi_han_boi_phan_trang(self, admin_client):
        """
        Bẫy dễ mắc nhất: danh sách phân trang 10 dòng/trang. Nếu export dùng
        nhầm trang hiện tại thay vì toàn bộ queryset, người dùng có 25 khách
        hàng sẽ chỉ xuất ra được 10 mà không hề biết.
        """
        for i in range(25):
            baker.make("projects.Client", client_name=f"Client {i:02d}")

        response = admin_client.get("/api/admin/clients/export/")
        sheet = doc_workbook(response).active
        assert sheet.max_row == 26  # 1 tieu de + 25 client

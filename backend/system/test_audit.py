import pytest
from system.models import AuditLog


# Kiểm thử API xem Audit Log tại /api/admin/audit-logs/ (read-only)
@pytest.mark.django_db
class TestAuditLogAPI:

    # GET danh sách → expect 200 và có ít nhất 1 log
    # /api/admin/audit-logs/ có pagination nên response.data là {count, results, ...}.
    def test_list_audit_logs(self, auth_client, admin_user):
        AuditLog.objects.create(
            user=admin_user, action='CREATE', table_name='clients', record_id=1
        )
        response = auth_client.get('/api/admin/audit-logs/')
        assert response.status_code == 200
        assert len(response.data['results']) >= 1

    # GET ?action=CREATE → expect chỉ trả log có action=CREATE
    def test_filter_by_action(self, auth_client, admin_user):
        AuditLog.objects.create(
            user=admin_user, action='CREATE', table_name='clients', record_id=1
        )
        AuditLog.objects.create(
            user=admin_user, action='DELETE', table_name='clients', record_id=2
        )
        response = auth_client.get('/api/admin/audit-logs/?action=CREATE')
        assert response.status_code == 200
        assert all(log['action'] == 'CREATE' for log in response.data['results'])

    # GET ?table_name=clients → expect chỉ trả log của bảng clients
    def test_filter_by_table_name(self, auth_client, admin_user):
        AuditLog.objects.create(
            user=admin_user, action='CREATE', table_name='clients', record_id=1
        )
        AuditLog.objects.create(
            user=admin_user, action='CREATE', table_name='jobs', record_id=1
        )
        response = auth_client.get('/api/admin/audit-logs/?table_name=clients')
        assert response.status_code == 200
        assert all(log['table_name'] == 'clients' for log in response.data['results'])

    # GET ?keyword=XYZ → expect chỉ trả log có chứa XYZ trong old/new values
    def test_filter_by_keyword(self, auth_client, admin_user):
        AuditLog.objects.create(
            user=admin_user, action='UPDATE', table_name='clients', record_id=1,
            new_values={'client_name': 'Cong ty XYZ'},
        )
        AuditLog.objects.create(
            user=admin_user, action='UPDATE', table_name='clients', record_id=2,
            new_values={'client_name': 'Cong ty ABC'},
        )
        response = auth_client.get('/api/admin/audit-logs/?keyword=XYZ')
        assert response.status_code == 200
        assert len(response.data['results']) == 1

    # GET ?record_id=99 → expect chỉ trả log của đúng record_id đó
    def test_filter_by_record_id(self, auth_client, admin_user):
        AuditLog.objects.create(
            user=admin_user, action='UPDATE', table_name='clients', record_id=99
        )
        AuditLog.objects.create(
            user=admin_user, action='UPDATE', table_name='clients', record_id=100
        )
        response = auth_client.get('/api/admin/audit-logs/?record_id=99')
        assert response.status_code == 200
        assert all(log['record_id'] == 99 for log in response.data['results'])

    # POST → expect 405 Method Not Allowed (AuditLogViewSet là ReadOnly)
    # AuditLog chỉ được đọc, không được tạo/sửa/xóa qua API
    def test_audit_log_readonly(self, auth_client):
        response = auth_client.post('/api/admin/audit-logs/', {'action': 'TEST'})
        assert response.status_code == 405


# Kiểm thử API Dashboard tại /api/admin/dashboard/
@pytest.mark.django_db
class TestDashboardAPI:

    # GET → expect 200
    def test_dashboard_returns_200(self, auth_client):
        response = auth_client.get('/api/admin/dashboard/')
        assert response.status_code == 200

    # GET → expect response có đủ key theo DashboardView hiện tại
    # (đã đổi field set — xem system/admin/views.py::DashboardView.get()).
    def test_dashboard_has_required_keys(self, auth_client):
        response = auth_client.get('/api/admin/dashboard/')
        data = response.data
        required_keys = [
            'active_clients', 'total_users', 'active_accounts', 'locked_accounts',
            'departments_without_manager', 'jobs_by_status', 'clients_overview',
            'audit_summary_today', 'recent_security_events',
        ]
        for key in required_keys:
            assert key in data, f"Missing key: {key}"

    # GET sau khi tạo client → expect active_clients đếm đúng
    def test_dashboard_counts_active_clients(self, auth_client):
        from projects.models import Client
        Client.objects.create(client_name='Active', tax_code='T001', is_active=True)
        Client.objects.create(client_name='Inactive', tax_code='T002', is_active=False)
        response = auth_client.get('/api/admin/dashboard/')
        assert response.data['active_clients'] >= 1
        assert response.data['clients_overview']['active'] >= 1


# Kiểm thử API xuất Excel riêng của từng tab Admin — endpoint gộp
# /api/admin/reports/ cũ đã được thay bằng action export trên từng ViewSet
XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


@pytest.mark.django_db
class TestExportAPI:

    # GET → expect 200, Content-Type là xlsx, Content-Disposition có filename
    def test_client_export_returns_excel_file(self, auth_client):
        response = auth_client.get('/api/admin/clients/export/')
        assert response.status_code == 200
        assert response['Content-Type'] == XLSX_CONTENT_TYPE
        assert 'attachment' in response['Content-Disposition']
        assert 'worktracker_clients.xlsx' in response['Content-Disposition']

    # GET → expect AuditLog được tạo với action=EXPORT, table_name=clients
    def test_client_export_creates_audit_log(self, auth_client):
        auth_client.get('/api/admin/clients/export/')
        assert AuditLog.objects.filter(
            action='EXPORT', table_name='clients', record_id=0
        ).exists()

    # Export tôn trọng đúng bộ lọc đang áp dụng, không dump toàn bộ bảng
    def test_client_export_respects_filters(self, auth_client):
        from openpyxl import load_workbook
        from io import BytesIO
        from projects.models import Client

        Client.objects.create(client_name='Active A', tax_code='E001', is_active=True)
        Client.objects.create(client_name='Inactive B', tax_code='E002', is_active=False)

        response = auth_client.get('/api/admin/clients/export/', {'is_active': 'true'})
        sheet = load_workbook(BytesIO(response.content)).active
        names = [sheet.cell(row=r, column=2).value for r in range(2, sheet.max_row + 1)]
        assert 'Active A' in names
        assert 'Inactive B' not in names

    # Các tab còn lại cũng phải trả về file xlsx hợp lệ
    def test_other_exports_return_excel(self, auth_client):
        for url in [
            '/api/admin/jobs/export/',
            '/api/auth/users/export/',
            '/api/auth/departments/export/',
            '/api/admin/audit-logs/export/',
        ]:
            response = auth_client.get(url)
            assert response.status_code == 200, url
            assert response['Content-Type'] == XLSX_CONTENT_TYPE, url

    # Endpoint gộp cũ đã bị gỡ bỏ → expect 404
    def test_old_combined_report_endpoint_removed(self, auth_client):
        response = auth_client.get('/api/admin/reports/')
        assert response.status_code == 404

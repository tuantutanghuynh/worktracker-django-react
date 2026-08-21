from io import BytesIO
from datetime import datetime
from django.template.loader import render_to_string
from xhtml2pdf import pisa

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

from rest_framework.exceptions import APIException

from reports.services.manager_task_summary_report_service import (
    build_task_summary_report,
)
from reports.services.manager_timesheet_detail_report_service import (
    build_timesheet_detail_report,
)

from system.services.audit_manager_service import log_action


class ReportExportError(APIException):
    status_code = 400
    default_detail = "Report export failed."
    default_code = "report_export_error"


def safe_value(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        if hasattr(value, "hour"):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    return value


def build_filename(report_type, file_format):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    extension = file_format.lower()
    return f"WorkTracker_{report_type.upper()}_{timestamp}.{extension}"


def get_report_data(*, user, filters):
    report_type = filters.get("report_type")

    if report_type == "TASK_SUMMARY":
        return build_task_summary_report(
            user=user,
            filters=filters,
        )

    if report_type in ["TIMESHEET_DETAIL", "TIMESHEET_EFFORT"]:
        return build_timesheet_detail_report(
            user=user,
            filters=filters,
        )

    raise ReportExportError("UNSUPPORTED_REPORT_TYPE")


# ============================================================
# EXCEL STYLING PALETTES & HELPERS
# ============================================================
NAVY_HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
SUB_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# KPI Card Fills
KPI_BLUE_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
KPI_GREEN_FILL = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
KPI_AMBER_FILL = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
KPI_ROSE_FILL = PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid")

# Status Fills
STATUS_GREEN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
STATUS_BLUE_FILL = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
STATUS_PURPLE_FILL = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
STATUS_AMBER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
STATUS_ROSE_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

STATUS_GREEN_FONT = Font(name="Calibri", size=10, bold=True, color="15803D")
STATUS_BLUE_FONT = Font(name="Calibri", size=10, bold=True, color="0369A1")
STATUS_PURPLE_FONT = Font(name="Calibri", size=10, bold=True, color="7E22CE")
STATUS_AMBER_FONT = Font(name="Calibri", size=10, bold=True, color="92400E")
STATUS_ROSE_FONT = Font(name="Calibri", size=10, bold=True, color="B91C1C")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUB_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="94A3B8")
REGULAR_FONT = Font(name="Calibri", size=10, color="1E293B")
BOLD_FONT = Font(name="Calibri", size=10, bold=True, color="0F172A")
CODE_FONT = Font(name="Consolas", size=10, bold=True, color="1E40AF")

THIN_BORDER_SIDE = Side(border_style="thin", color="CBD5E1")
THIN_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def style_table_header(row_cells):
    for cell in row_cells:
        cell.fill = NAVY_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def apply_zebra_and_borders(sheet, start_row, max_row, max_col):
    for r in range(start_row, max_row + 1):
        is_even = (r % 2 == 0)
        row_fill = ZEBRA_FILL if is_even else WHITE_FILL
        for c in range(1, max_col + 1):
            cell = sheet.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if cell.fill == WHITE_FILL or cell.fill.fill_type is None:
                cell.fill = row_fill


def autosize_worksheet_columns(workbook):
    for sheet in workbook.worksheets:
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                val = str(cell.value or "")
                if val:
                    lines = val.split("\n")
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)

            adjusted_width = max(max_length + 3, 11)
            # Giới hạn tối đa 45 để bảng không bị quá bè
            sheet.column_dimensions[col_letter].width = min(adjusted_width, 45)


# ============================================================
# 1. TASK SUMMARY EXCEL BUILDER
# ============================================================
def write_task_summary_sheet(workbook, report_data):
    # --- TAB 1: SUMMARY & ANALYTICS DASHBOARD ---
    sum_sheet = workbook.active
    sum_sheet.title = "Executive Summary"
    sum_sheet.views.sheetView[0].showGridLines = True

    # 1. Top Corporate Banner
    sum_sheet.merge_cells("A1:H2")
    banner_cell = sum_sheet["A1"]
    banner_cell.value = "WORKTRACKER PRO  •  EXECUTIVE TASK DELIVERY REPORT"
    banner_cell.fill = NAVY_HEADER_FILL
    banner_cell.font = TITLE_FONT
    banner_cell.alignment = ALIGN_CENTER

    sum_sheet.merge_cells("A3:H3")
    sub_banner = sum_sheet["A3"]
    sub_banner.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Scope: Manager Managed Projects"
    sub_banner.fill = SUB_HEADER_FILL
    sub_banner.font = SUBTITLE_FONT
    sub_banner.alignment = ALIGN_CENTER

    # 2. KPI Cards (Row 5 - 6)
    kpis = [
        ("TOTAL TASKS", report_data["summary"]["total_tasks"], KPI_BLUE_FILL, "1D4ED8", "A", "B"),
        ("COMPLETED TASKS", report_data["summary"]["status_summary"].get("COMPLETED", 0), KPI_GREEN_FILL, "047857", "C", "D"),
        ("ACTIVE IN-PROGRESS", report_data["summary"]["overdue_summary"].get("total_active_tasks", 0), KPI_AMBER_FILL, "B45309", "E", "F"),
        (f"OVERDUE ({report_data['summary']['overdue_summary'].get('overdue_rate_percent', 0)}%)", report_data["summary"]["overdue_summary"].get("overdue_tasks", 0), KPI_ROSE_FILL, "BE123C", "G", "H"),
    ]

    for title, val, fill, color_hex, col_start, col_end in kpis:
        sum_sheet.merge_cells(f"{col_start}5:{col_end}5")
        sum_sheet.merge_cells(f"{col_start}6:{col_end}6")

        c_title = sum_sheet[f"{col_start}5"]
        c_title.value = title
        c_title.fill = fill
        c_title.font = Font(name="Calibri", size=9, bold=True, color="475569")
        c_title.alignment = ALIGN_CENTER
        c_title.border = THIN_BORDER

        c_val = sum_sheet[f"{col_start}6"]
        c_val.value = val
        c_val.fill = fill
        c_val.font = Font(name="Calibri", size=16, bold=True, color=color_hex)
        c_val.alignment = ALIGN_CENTER
        c_val.border = THIN_BORDER

        # Border for merged pair
        for row in range(5, 7):
            for col_idx in [get_column_letter(sum_sheet[f"{col_end}5"].column)]:
                sum_sheet[f"{col_idx}{row}"].border = THIN_BORDER

    # 3. Status Breakdown Table (Row 8 - 14)
    sum_sheet["A8"] = "Task Status Breakdown"
    sum_sheet.merge_cells("A8:C8")
    sum_sheet["A8"].fill = SUB_HEADER_FILL
    sum_sheet["A8"].font = SUB_HEADER_FONT
    sum_sheet["A8"].alignment = ALIGN_CENTER

    sum_sheet["A9"] = "Status"
    sum_sheet["B9"] = "Tasks Count"
    sum_sheet["C9"] = "Proportion"
    style_table_header([sum_sheet["A9"], sum_sheet["B9"], sum_sheet["C9"]])

    total_tasks = report_data["summary"]["total_tasks"] or 1
    curr_row = 10
    status_summary = report_data["summary"]["status_summary"]
    status_order = ["TODO", "IN_PROGRESS", "REVIEWING", "COMPLETED", "CANCELLED"]

    for st in status_order:
        cnt = status_summary.get(st, 0)
        sum_sheet[f"A{curr_row}"] = st
        sum_sheet[f"B{curr_row}"] = cnt
        sum_sheet[f"C{curr_row}"] = f"{(cnt / total_tasks * 100):.1f}%"

        sum_sheet[f"A{curr_row}"].font = BOLD_FONT
        sum_sheet[f"A{curr_row}"].alignment = ALIGN_LEFT
        sum_sheet[f"B{curr_row}"].font = REGULAR_FONT
        sum_sheet[f"B{curr_row}"].alignment = ALIGN_CENTER
        sum_sheet[f"C{curr_row}"].font = REGULAR_FONT
        sum_sheet[f"C{curr_row}"].alignment = ALIGN_RIGHT
        curr_row += 1

    apply_zebra_and_borders(sum_sheet, 10, curr_row - 1, 3)

    # 4. Native Excel Chart (Bar Chart)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Task Distribution by Status"
    chart.y_axis.title = "Number of Tasks"
    chart.x_axis.title = "Status"
    chart.height = 8.5
    chart.width = 14

    data_ref = Reference(sum_sheet, min_col=2, min_row=9, max_row=curr_row - 1)
    cats_ref = Reference(sum_sheet, min_col=1, min_row=10, max_row=curr_row - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None

    sum_sheet.add_chart(chart, "E8")

    # --- TAB 2: DETAILED TASK REGISTER ---
    data_sheet = workbook.create_sheet("Task Details")
    data_sheet.views.sheetView[0].showGridLines = True

    headers = [
        "Task ID",
        "Task Title",
        "Job Code",
        "Project (Job) Name",
        "Assignee",
        "Priority",
        "Status",
        "Deadline",
        "Completed At",
        "Created At",
        "Updated At",
    ]
    data_sheet.append(headers)
    style_table_header(data_sheet[1])
    data_sheet.row_dimensions[1].height = 25

    row_idx = 2
    for row in report_data["rows"]:
        st = row.get("status", "")
        pr = row.get("priority", "")

        data_sheet.append(
            [
                f"TSK-{row.get('id')}",
                safe_value(row.get("title")),
                safe_value(row.get("job", {}).get("job_code")),
                safe_value(row.get("job", {}).get("job_name")),
                safe_value(row.get("assignee", {}).get("full_name")),
                pr,
                st,
                safe_value(row.get("deadline")),
                safe_value(row.get("completed_at")),
                safe_value(row.get("created_at")),
                safe_value(row.get("updated_at")),
            ]
        )

        # Style individual cells
        data_sheet.cell(row=row_idx, column=1).font = CODE_FONT
        data_sheet.cell(row=row_idx, column=1).alignment = ALIGN_CENTER
        data_sheet.cell(row=row_idx, column=2).font = BOLD_FONT

        # Status badge colors
        st_cell = data_sheet.cell(row=row_idx, column=7)
        st_cell.alignment = ALIGN_CENTER
        if st == "COMPLETED":
            st_cell.fill, st_cell.font = STATUS_GREEN_FILL, STATUS_GREEN_FONT
        elif st == "IN_PROGRESS":
            st_cell.fill, st_cell.font = STATUS_BLUE_FILL, STATUS_BLUE_FONT
        elif st == "REVIEWING":
            st_cell.fill, st_cell.font = STATUS_PURPLE_FILL, STATUS_PURPLE_FONT
        elif st == "TODO":
            st_cell.fill, st_cell.font = STATUS_AMBER_FILL, STATUS_AMBER_FONT
        elif st == "CANCELLED":
            st_cell.fill, st_cell.font = STATUS_ROSE_FILL, STATUS_ROSE_FONT

        row_idx += 1

    apply_zebra_and_borders(data_sheet, 2, row_idx - 1, len(headers))
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"


# ============================================================
# 2. TIMESHEET DETAIL EXCEL BUILDER
# ============================================================
def write_timesheet_detail_sheet(workbook, report_data):
    # --- TAB 1: EXECUTIVE SUMMARY & CHARTS ---
    sum_sheet = workbook.active
    sum_sheet.title = "Executive Summary"
    sum_sheet.views.sheetView[0].showGridLines = True

    # 1. Top Corporate Banner
    sum_sheet.merge_cells("A1:H2")
    banner_cell = sum_sheet["A1"]
    banner_cell.value = "WORKTRACKER PRO  •  TIMESHEET & EFFORT AUDIT REPORT"
    banner_cell.fill = NAVY_HEADER_FILL
    banner_cell.font = TITLE_FONT
    banner_cell.alignment = ALIGN_CENTER

    sum_sheet.merge_cells("A3:H3")
    sub_banner = sum_sheet["A3"]
    sub_banner.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Scope: Manager Managed Projects"
    sub_banner.fill = SUB_HEADER_FILL
    sub_banner.font = SUBTITLE_FONT
    sub_banner.alignment = ALIGN_CENTER

    # 2. KPI Cards (Row 5 - 6)
    kpis = [
        ("TOTAL LOGS", report_data["summary"]["total_logs"], KPI_BLUE_FILL, "1D4ED8", "A", "B"),
        ("TOTAL HOURS", f"{report_data['summary']['total_hours']} hrs", KPI_GREEN_FILL, "047857", "C", "D"),
        ("APPROVED HOURS", f"{report_data['summary'].get('approved_hours', report_data['summary'].get('total_hours', 0))} hrs", KPI_AMBER_FILL, "B45309", "E", "F"),
        ("PENDING / REJECTED", f"{report_data['summary']['review_status_summary'].get('PENDING', 0) + report_data['summary']['review_status_summary'].get('REJECTED', 0)} logs", KPI_ROSE_FILL, "BE123C", "G", "H"),
    ]

    for title, val, fill, color_hex, col_start, col_end in kpis:
        sum_sheet.merge_cells(f"{col_start}5:{col_end}5")
        sum_sheet.merge_cells(f"{col_start}6:{col_end}6")

        c_title = sum_sheet[f"{col_start}5"]
        c_title.value = title
        c_title.fill = fill
        c_title.font = Font(name="Calibri", size=9, bold=True, color="475569")
        c_title.alignment = ALIGN_CENTER
        c_title.border = THIN_BORDER

        c_val = sum_sheet[f"{col_start}6"]
        c_val.value = str(val)
        c_val.fill = fill
        c_val.font = Font(name="Calibri", size=16, bold=True, color=color_hex)
        c_val.alignment = ALIGN_CENTER
        c_val.border = THIN_BORDER

        for row in range(5, 7):
            for col_idx in [get_column_letter(sum_sheet[f"{col_end}5"].column)]:
                sum_sheet[f"{col_idx}{row}"].border = THIN_BORDER

    # 3. Effort Distribution by Project (Row 8 - 15)
    sum_sheet["A8"] = "Effort by Project"
    sum_sheet.merge_cells("A8:C8")
    sum_sheet["A8"].fill = SUB_HEADER_FILL
    sum_sheet["A8"].font = SUB_HEADER_FONT
    sum_sheet["A8"].alignment = ALIGN_CENTER

    sum_sheet["A9"] = "Project Name"
    sum_sheet["B9"] = "Logs Count"
    sum_sheet["C9"] = "Total Hours"
    style_table_header([sum_sheet["A9"], sum_sheet["B9"], sum_sheet["C9"]])

    curr_row = 10
    job_summary = report_data["summary"].get("job_summary", [])
    for job_row in job_summary:
        sum_sheet[f"A{curr_row}"] = job_row.get("job_name", "Project")
        sum_sheet[f"B{curr_row}"] = job_row.get("total_logs", 0)
        sum_sheet[f"C{curr_row}"] = float(job_row.get("total_hours", 0))

        sum_sheet[f"A{curr_row}"].font = BOLD_FONT
        sum_sheet[f"B{curr_row}"].font = REGULAR_FONT
        sum_sheet[f"B{curr_row}"].alignment = ALIGN_CENTER
        sum_sheet[f"C{curr_row}"].font = REGULAR_FONT
        sum_sheet[f"C{curr_row}"].alignment = ALIGN_RIGHT
        sum_sheet[f"C{curr_row}"].number_format = "#,##0.0"
        curr_row += 1

    if curr_row > 10:
        apply_zebra_and_borders(sum_sheet, 10, curr_row - 1, 3)

        # Embedded Pie Chart for Project Effort
        pie = PieChart()
        pie.title = "Hours by Project"
        labels = Reference(sum_sheet, min_col=1, min_row=10, max_row=curr_row - 1)
        data = Reference(sum_sheet, min_col=3, min_row=9, max_row=curr_row - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.style = 10
        pie.height = 8.5
        pie.width = 14
        sum_sheet.add_chart(pie, "E8")

    # --- TAB 2: DETAILED TIMESHEET REGISTER ---
    data_sheet = workbook.create_sheet("Timesheet Details")
    data_sheet.views.sheetView[0].showGridLines = True

    headers = [
        "Log ID",
        "Work Date",
        "Employee Name",
        "Department",
        "Job Code",
        "Project Name",
        "Task Title",
        "Task Status",
        "Hours",
        "Description",
        "Review Status",
        "Reviewed By",
        "Reviewed At",
        "Review Note",
        "Adjusted By",
        "Adjusted At",
        "Adjustment Reason",
        "Period Status",
        "Created At",
    ]
    data_sheet.append(headers)
    style_table_header(data_sheet[1])
    data_sheet.row_dimensions[1].height = 25

    row_idx = 2
    for row in report_data["rows"]:
        employee = row.get("user") or row.get("employee") or {}
        department = employee.get("department")
        reviewed_by = row.get("reviewed_by") or {}
        adjusted_by = row.get("adjusted_by") or {}
        rv_status = row.get("review_status", "")

        data_sheet.append(
            [
                f"LW-{row.get('id')}",
                safe_value(row.get("work_date")),
                safe_value(employee.get("full_name") if isinstance(employee, dict) else ""),
                safe_value(department.get("name") if isinstance(department, dict) else ""),
                safe_value(row.get("job", {}).get("job_code")),
                safe_value(row.get("job", {}).get("job_name")),
                safe_value(row.get("task", {}).get("title")),
                safe_value(row.get("task", {}).get("status")),
                float(row.get("hours_spent", 0)),
                safe_value(row.get("description")),
                rv_status,
                safe_value(reviewed_by.get("full_name") if isinstance(reviewed_by, dict) else ""),
                safe_value(row.get("reviewed_at")),
                safe_value(row.get("review_note")),
                safe_value(adjusted_by.get("full_name") if isinstance(adjusted_by, dict) else ""),
                safe_value(row.get("adjusted_at")),
                safe_value(row.get("adjustment_reason")),
                safe_value(row.get("locked_period_status")),
                safe_value(row.get("created_at")),
            ]
        )

        # Style cells
        data_sheet.cell(row=row_idx, column=1).font = CODE_FONT
        data_sheet.cell(row=row_idx, column=1).alignment = ALIGN_CENTER
        data_sheet.cell(row=row_idx, column=2).alignment = ALIGN_CENTER

        # Hours format
        h_cell = data_sheet.cell(row=row_idx, column=9)
        h_cell.font = BOLD_FONT
        h_cell.alignment = ALIGN_RIGHT
        h_cell.number_format = "#,##0.0"

        # Review status styling
        rv_cell = data_sheet.cell(row=row_idx, column=11)
        rv_cell.alignment = ALIGN_CENTER
        if rv_status == "APPROVED":
            rv_cell.fill, rv_cell.font = STATUS_GREEN_FILL, STATUS_GREEN_FONT
        elif rv_status == "PENDING":
            rv_cell.fill, rv_cell.font = STATUS_AMBER_FILL, STATUS_AMBER_FONT
        elif rv_status == "REJECTED":
            rv_cell.fill, rv_cell.font = STATUS_ROSE_FILL, STATUS_ROSE_FONT

        row_idx += 1

    apply_zebra_and_borders(data_sheet, 2, row_idx - 1, len(headers))
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"


# ============================================================
# EXPORT DISPATCHER
# ============================================================
def export_xlsx(*, report_type, report_data):
    workbook = Workbook()

    if report_type == "TASK_SUMMARY":
        write_task_summary_sheet(workbook, report_data)
    elif report_type in ["TIMESHEET_DETAIL", "TIMESHEET_EFFORT"]:
        write_timesheet_detail_sheet(workbook, report_data)
    else:
        raise ReportExportError("UNSUPPORTED_REPORT_TYPE")

    autosize_worksheet_columns(workbook)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_pdf(*, report_type, report_data):
    if report_type == "TASK_SUMMARY":
        template_path = "reports/task_summary_pdf.html"
    elif report_type in ["TIMESHEET_DETAIL", "TIMESHEET_EFFORT"]:
        template_path = "reports/timesheet_detail_pdf.html"
    else:
        raise ReportExportError("UNSUPPORTED_REPORT_TYPE")

    html_string = render_to_string(template_path, report_data)

    output = BytesIO()
    pisa_status = pisa.CreatePDF(
        src=html_string,
        dest=output,
        encoding="utf-8",
    )

    if pisa_status.err:
        raise ReportExportError("Error during PDF generation.")

    return output.getvalue()


def export_manager_report(*, user, filters, request=None):
    report_type = filters.get("report_type")
    file_format = filters.get("file_format")

    report_data = get_report_data(
        user=user,
        filters=filters,
    )

    if file_format == "XLSX":
        content = export_xlsx(
            report_type=report_type,
            report_data=report_data,
        )
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    elif file_format == "PDF":
        content = export_pdf(
            report_type=report_type,
            report_data=report_data,
        )
        content_type = "application/pdf"
    else:
        raise ReportExportError("UNSUPPORTED_EXPORT_FORMAT")

    filename = build_filename(
        report_type=report_type,
        file_format=file_format,
    )

    log_action(
        user=user,
        action="REPORT_EXPORTED",
        table_name="reports",
        record_id=0,
        old_values=None,
        new_values={
            "report_type": report_type,
            "file_format": file_format,
            "filename": filename,
            "filters": dict(filters),
        },
        summary=f"Manager exported {report_type} report as {file_format} file ({filename})",
        request=request,
    )

    return {
        "filename": filename,
        "content_type": content_type,
        "content": content,
    }